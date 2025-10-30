import pdfplumber
import sys
import os
# Streamlit 환경에서는 tkinter를 사용하지 않음
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    TKINTER_AVAILABLE = True
except ImportError:
    TKINTER_AVAILABLE = False
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import re
import subprocess
import platform
import threading
import time
import json

def get_config_file_path():
    """
    설정 파일 경로를 반환하는 함수
    
    Returns:
        str: 설정 파일의 전체 경로
    """
    # 현재 스크립트와 같은 디렉토리에 config.json 파일 생성
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(script_dir, "pdf_converter_config.json")

def load_last_directory():
    """
    마지막으로 사용한 디렉토리를 불러오는 함수
    
    Returns:
        str: 마지막 사용 디렉토리 경로, 없으면 현재 디렉토리
    """
    config_file = get_config_file_path()
    try:
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                last_dir = config.get('last_pdf_directory', os.getcwd())
                # 디렉토리가 실제로 존재하는지 확인
                if os.path.exists(last_dir):
                    return last_dir
    except (json.JSONDecodeError, FileNotFoundError, PermissionError):
        pass
    
    return os.getcwd()

def save_last_directory(directory_path):
    """
    마지막으로 사용한 디렉토리를 저장하는 함수
    
    Args:
        directory_path (str): 저장할 디렉토리 경로
    """
    config_file = get_config_file_path()
    try:
        # 기존 설정이 있으면 불러오기
        config = {}
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            except (json.JSONDecodeError, FileNotFoundError):
                config = {}
        
        # 마지막 디렉토리 업데이트
        config['last_pdf_directory'] = directory_path
        
        # 설정 파일에 저장
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except (PermissionError, OSError):
        # 저장 실패시 무시 (치명적이지 않음)
        pass

class ProgressWindow:
    """
    프로그래스바를 표시하는 GUI 클래스
    """
    def __init__(self):
        if not TKINTER_AVAILABLE:
            return
        
        self.root = tk.Tk()
        self.root.title("PDF 처리 중...")
        
        # 창 크기 설정
        window_width = 400
        window_height = 150
        
        # 화면 크기 가져오기
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # 화면 중앙 위치 계산
        center_x = int(screen_width/2 - window_width/2)
        center_y = int(screen_height/2 - window_height/2)
        
        # 창 크기와 위치 설정
        self.root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
        self.root.resizable(False, False)
        self.root.attributes('-topmost', True)  # 항상 위에 표시
        
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 제목 라벨
        self.title_label = ttk.Label(main_frame, text="Processing PDF file...", 
                                    font=("Arial", 12, "bold"))
        self.title_label.pack(pady=(0, 20))
        
        # 상태 라벨
        self.status_label = ttk.Label(main_frame, text="Starting...", 
                                     font=("Arial", 10),
                                     width=50,
                                     anchor="center")
        self.status_label.pack(pady=(0, 10))
        
        # 프로그래스바
        self.progress = ttk.Progressbar(main_frame, length=300, mode='determinate')
        self.progress.pack(pady=(0, 10))
        
        # 퍼센트 라벨
        self.percent_label = ttk.Label(main_frame, text="0%", 
                                      font=("Arial", 10),
                                      width=10,
                                      anchor="center")
        self.percent_label.pack()
        
        # 프로그래스바 초기화
        self.progress['maximum'] = 100
        self.progress['value'] = 0
        
    def update_progress(self, value, status_text=""):
        """
        프로그래스바 업데이트
        
        Args:
            value (int): 프로그래스 값 (0-100)
            status_text (str): 상태 텍스트
        """
        if not TKINTER_AVAILABLE:
            return
        
        self.progress['value'] = value
        self.percent_label.config(text=f"{value}%")
        if status_text:
            # 이전 텍스트를 완전히 지우고 새 텍스트로 교체
            self.status_label.config(text="")
            self.root.update_idletasks()
            self.status_label.config(text=status_text)
        self.root.update_idletasks()
        
    def close(self):
        """프로그래스바 창 닫기"""
        if not TKINTER_AVAILABLE:
            return
        
        self.root.destroy()
        
    def show(self):
        """프로그래스바 창 표시"""
        if not TKINTER_AVAILABLE:
            return
        
        self.root.update()

def extract_data_from_first_page(lines):
    """
    첫 번째 페이지의 특정 줄에서 데이터를 추출하는 함수
    8번째 줄에서 Sample ID와 Date 추출, 13~30번째 줄에서 데이터 처리
    
    Args:
        lines (list): 페이지의 모든 줄들
        
    Returns:
        tuple: (sample_id, date, extracted_data)
    """
    sample_id = None
    date = None
    extracted_data = []
    current_row_data = {}
    
    # 숫자형 변환 함수
    def convert_to_number(text):
        """텍스트를 적절한 숫자형으로 변환"""
        try:
            normalized = text.replace(',', '.')
            if '.' in normalized:
                return float(normalized)
            else:
                return int(normalized)
        except ValueError:
            return text  # 변환 실패시 원본 반환
    
    # 8번째 줄에서 Sample ID와 Date 추출 (인덱스 7)
    if len(lines) > 7:
        line_8 = lines[7].strip()
        
        # SerumPlasma 또는 Ser/PI 패턴으로 Sample ID와 Date 추출
        if "SerumPlasma" in line_8 or "Ser/PI" in line_8:
            parts = line_8.split()
            
            if "SerumPlasma" in line_8:
                # SerumPlasma 형태: SerumPlasma 50016-1 ID : 187 Test Sample 2023/12/08 19:23:06
                # "ID :"와 "YYYY/MM/DD" 사이의 모든 단어를 Sample ID로 추출
                id_colon_found = False
                sample_id_parts = []
                
                for i, part in enumerate(parts):
                    # "ID :" 패턴 찾기
                    if part == "ID" and i + 1 < len(parts) and parts[i + 1] == ":":
                        id_colon_found = True
                        start_idx = i + 2  # "ID :" 다음부터 시작
                        
                        # "ID :" 다음부터 날짜 패턴 전까지 모든 단어 수집
                        for j in range(start_idx, len(parts)):
                            current_part = parts[j]
                            # 날짜 패턴(YYYY/MM/DD)이 나오면 중단
                            if re.match(r'\d{4}/\d{2}/\d{2}', current_part):
                                break
                            sample_id_parts.append(current_part)
                        break
                
                # Sample ID 완성 (공백으로 연결)
                if sample_id_parts:
                    sample_id = " ".join(sample_id_parts)
                    
            elif "Ser/PI" in line_8:
                # Ser/PI 형태: Ser/PI 50016-1 2023/12/08 19:23:06
                # 두번째 단어를 Sample ID로 추출 (기존 로직 유지)
                if len(parts) >= 2:
                    sample_id = parts[1]  # "Ser/PI" 다음 단어
            
            # 날짜는 같은 줄에서 YYYY/MM/DD 형태로 찾기
            for part in parts:
                if re.match(r'\d{4}/\d{2}/\d{2}', part):
                    date = part
                    break
    
    # 13번째 줄부터 30번째 줄까지 처리 (인덱스 12부터 29까지)
    start_line = 12  # 13번째 줄 (0-based index)
    end_line = min(30, len(lines))  # 30번째 줄까지 또는 페이지 끝까지
    
    i = start_line
    while i < end_line:
        if i >= len(lines):
            break
            
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        
        parts = line.split()
        if not parts:
            i += 1
            continue
        
        # "R2" 또는 "R3"로 시작하는 줄은 생략
        if parts[0] in ["R2", "R3"]:
            i += 1
            continue
        
        # Test Name과 Result 패턴 처리 (개선된 정규식 패턴)
        # 정규식 패턴: +로 시작할 수 있고, 대문자로 시작하는 영문/숫자/하이픈 조합
        if re.match(r'^[\+]?[A-Z][A-Z0-9\-]+', line) or line.startswith('+'):
            # "+" 존재 여부 확인
            has_plus = line.startswith('+')
            
            # 일반적인 테스트 라인 처리
            if has_plus:
                # "+" 있는 경우: 두번째 단어부터 처리
                start_idx = 1
                rerun = "Y"
            else:
                # "+" 없는 경우: 첫번째 단어부터 처리
                start_idx = 0
                rerun = "N"
            
            # Test Name 구성: 첫 단어 + 1자리 숫자나 v2/V2 패턴 단어들
            test_name_parts = []
            current_idx = start_idx
            
            # 첫 번째 테스트명 단어 추가
            if current_idx < len(parts):
                test_name_parts.append(parts[current_idx])
                current_idx += 1
            
            # 다음 단어들 중 1자리 숫자나 v숫자/V숫자 패턴이면 Test Name에 포함
            while current_idx < len(parts):
                word = parts[current_idx]
                # 1자리 숫자이거나 v숫자/V숫자 형태인 경우만 Test Name에 포함
                if ((word.isdigit() and len(word) == 1) or 
                    (word.lower().startswith('v') and len(word) > 1 and word[1:].isdigit())):
                    test_name_parts.append(word)
                    current_idx += 1
                else:
                    break
            
            # Test Name 완성
            test_name = " ".join(test_name_parts)
            
            # Result 찾기: Test Name 다음의 첫 번째 숫자형 값
            result = ""
            result_idx = current_idx
            data_alarm = "N"
            
            # 숫자형 문자열을 찾는 함수
            def is_numeric(text):
                try:
                    # "4,12" -> "4.12"로 변환 후 숫자 확인
                    normalized = text.replace(',', '.')
                    float(normalized)
                    return True
                except ValueError:
                    return False
            
            # Result 추출
            while result_idx < len(parts):
                word = parts[result_idx]
                if is_numeric(word):
                    result = word.replace(',', '.')  # "4,12" -> "4.12"
                    
                    # Data Alarm 판정: Result 뒤에 추가 단어가 있으면 'Y'
                    if result_idx + 1 < len(parts):
                        data_alarm = "Y"
                    break
                result_idx += 1
            
            current_row_data = {
                'test_name': test_name,
                'result': result,
                'data_alarm': data_alarm,
                'rerun': rerun,
                'has_rerun': has_plus
            }
            
            # 다음 줄에서 Unit, AU, R.P Lot 정보 추출
            i += 1
            if i < len(lines):
                next_line = lines[i].strip()
                next_parts = next_line.split()
                
                # 다음 줄의 첫 번째 단어를 Unit으로 사용
                if next_parts:
                    unit = next_parts[0]
                    
                    # E열(AU) 로직: 단위 다음 단어에 "-"가 없으면 같은 줄에서 "-" 포함된 단어 찾기
                    au = ""
                    
                    # 두번째 문장이 "NACL"인 경우 특수 처리
                    if len(next_parts) > 2 and next_parts[1] == "NACL":
                        # NACL인 경우: 세번째 단어 확인
                        if len(next_parts) > 2:
                            potential_au = next_parts[2]
                            if "-" in potential_au:
                                au = potential_au
                            else:
                                # 세번째 단어에 "-"가 없으면 같은 줄에서 "-" 포함된 단어 찾기
                                for word in next_parts:
                                    if "-" in word:
                                        au = word
                                        break
                        
                        rp_lot = ""
                        if len(next_parts) >= 6:  # 다섯번째 문장 (인덱스 4)
                            potential_rp_lot = next_parts[4]
                            # 숫자로만 구성되어 있거나 숫자가 포함된 문자열인 경우만 R.P Lot으로 인식
                            if potential_rp_lot.isdigit() or (potential_rp_lot.isalnum() and any(c.isdigit() for c in potential_rp_lot)):
                                rp_lot = potential_rp_lot
                    else:
                        # 일반적인 경우: 두번째 단어 확인
                        if len(next_parts) > 1:
                            potential_au = next_parts[1]
                            if "-" in potential_au:
                                au = potential_au
                            else:
                                # 두번째 단어에 "-"가 없으면 같은 줄에서 "-" 포함된 단어 찾기
                                for word in next_parts:
                                    if "-" in word:
                                        au = word
                                        break
                        
                        rp_lot = ""
                        if len(next_parts) >= 5 and len(next_parts) > 3:
                            potential_rp_lot = next_parts[3]
                            # 숫자로만 구성되어 있거나 숫자가 포함된 문자열인 경우만 R.P Lot으로 인식
                            if potential_rp_lot.isdigit() or (potential_rp_lot.isalnum() and any(c.isdigit() for c in potential_rp_lot)):
                                rp_lot = potential_rp_lot
                    
                    # COI인 경우 R/NR 값 추출
                    r_nr_value = ""
                    if unit == "COI":
                        # 다음 줄에서 Reac 또는 NonReac 찾기
                        if i + 1 < len(lines):
                            next_next_line = lines[i + 1].strip()
                            if "Reac" in next_next_line:
                                if "NonReac" in next_next_line:
                                    r_nr_value = "NonReac"
                                else:
                                    r_nr_value = "Reac"
                    
                    # 현재 행 데이터 완성
                    row_data = {
                        'sample_id': sample_id,
                        'test_name': current_row_data.get('test_name', ''),
                        'result': current_row_data.get('result', ''),
                        'unit': unit,
                        'au': au,
                        'rp_lot': rp_lot,
                        'data_alarm': current_row_data.get('data_alarm', 'N'),
                        'rerun': current_row_data.get('rerun', 'N'),
                        'date': date,
                        'has_rerun': current_row_data.get('has_rerun', False),
                        'r_nr': r_nr_value
                    }
                    extracted_data.append(row_data)
                    current_row_data = {}  # 다음 데이터를 위해 초기화
        
        i += 1
    
    return sample_id, date, extracted_data

def extract_data_from_other_pages(lines):
    """
    두 번째 페이지부터의 특정 줄에서 데이터를 추출하는 함수
    5번째 줄에서 Sample ID와 Date 추출, 10번째 줄부터 30번째 줄까지 데이터 처리
    
    Args:
        lines (list): 페이지의 모든 줄들
        
    Returns:
        tuple: (sample_id, date, extracted_data)
    """
    extracted_data = []
    current_row_data = {}
    sample_id = None
    date = None
    
    # 5번째 줄에서 Sample ID와 Date 추출 (인덱스 4)
    if len(lines) > 4:
        line_5 = lines[4].strip()
        
        # SerumPlasma 또는 Ser/PI 패턴으로 Sample ID와 Date 추출
        if "SerumPlasma" in line_5 or "Ser/PI" in line_5:
            parts = line_5.split()
            
            if "SerumPlasma" in line_5:
                # SerumPlasma 형태: SerumPlasma 50016-1 ID : 187 Test Sample 2023/12/08 19:23:06
                # "ID :"와 "YYYY/MM/DD" 사이의 모든 단어를 Sample ID로 추출
                id_colon_found = False
                sample_id_parts = []
                
                for i, part in enumerate(parts):
                    # "ID :" 패턴 찾기
                    if part == "ID" and i + 1 < len(parts) and parts[i + 1] == ":":
                        id_colon_found = True
                        start_idx = i + 2  # "ID :" 다음부터 시작
                        
                        # "ID :" 다음부터 날짜 패턴 전까지 모든 단어 수집
                        for j in range(start_idx, len(parts)):
                            current_part = parts[j]
                            # 날짜 패턴(YYYY/MM/DD)이 나오면 중단
                            if re.match(r'\d{4}/\d{2}/\d{2}', current_part):
                                break
                            sample_id_parts.append(current_part)
                        break
                
                # Sample ID 완성 (공백으로 연결)
                if sample_id_parts:
                    sample_id = " ".join(sample_id_parts)
                    
            elif "Ser/PI" in line_5:
                # Ser/PI 형태: Ser/PI 50016-1 2023/12/08 19:23:06
                # 두번째 단어를 Sample ID로 추출 (기존 로직 유지)
                if len(parts) >= 2:
                    sample_id = parts[1]  # "Ser/PI" 다음 단어
            
            # 날짜는 같은 줄에서 YYYY/MM/DD 형태로 찾기
            for part in parts:
                if re.match(r'\d{4}/\d{2}/\d{2}', part):
                    date = part
                    break
    
    # 10번째 줄부터 30번째 줄까지 처리 (인덱스 9부터 29까지)
    start_line = 9  # 10번째 줄 (0-based index)
    end_line = min(30, len(lines))  # 30번째 줄까지 또는 페이지 끝까지
    
    i = start_line
    while i < end_line:
        if i >= len(lines):
            break
            
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        
        parts = line.split()
        if not parts:
            i += 1
            continue
            
        # "R2" 또는 "R3"로 시작하는 줄은 생략
        if parts[0] in ["R2", "R3"]:
            i += 1
            continue
        
        # Test Name과 Result 패턴 처리 (개선된 정규식 패턴)
        # 정규식 패턴: +로 시작할 수 있고, 대문자로 시작하는 영문/숫자/하이픈 조합
        if re.match(r'^[\+]?[A-Z][A-Z0-9\-]+', line) or line.startswith('+'):
            # "+" 존재 여부 확인
            has_plus = line.startswith('+')
            
            # 일반적인 테스트 라인 처리
            if has_plus:
                # "+" 있는 경우: 두번째 단어부터 처리
                start_idx = 1
                rerun = "Y"
            else:
                # "+" 없는 경우: 첫번째 단어부터 처리
                start_idx = 0
                rerun = "N"
            
            # Test Name 구성: 첫 단어 + 1자리 숫자나 v2/V2 패턴 단어들
            test_name_parts = []
            current_idx = start_idx
            
            # 첫 번째 테스트명 단어 추가
            if current_idx < len(parts):
                test_name_parts.append(parts[current_idx])
                current_idx += 1
            
            # 다음 단어들 중 1자리 숫자나 v숫자/V숫자 패턴이면 Test Name에 포함
            while current_idx < len(parts):
                word = parts[current_idx]
                # 1자리 숫자이거나 v숫자/V숫자 형태인 경우만 Test Name에 포함
                if ((word.isdigit() and len(word) == 1) or 
                    (word.lower().startswith('v') and len(word) > 1 and word[1:].isdigit())):
                    test_name_parts.append(word)
                    current_idx += 1
                else:
                    break
            
            # Test Name 완성
            test_name = " ".join(test_name_parts)
            
            # Result 찾기: Test Name 다음의 첫 번째 숫자형 값
            result = ""
            result_idx = current_idx
            data_alarm = "N"
            
            # 숫자형 문자열을 찾는 함수
            def is_numeric(text):
                try:
                    # "4,12" -> "4.12"로 변환 후 숫자 확인
                    normalized = text.replace(',', '.')
                    float(normalized)
                    return True
                except ValueError:
                    return False
            
            # Result 추출
            while result_idx < len(parts):
                word = parts[result_idx]
                if is_numeric(word):
                    result = word.replace(',', '.')  # "4,12" -> "4.12"
                    
                    # Data Alarm 판정: Result 뒤에 추가 단어가 있으면 'Y'
                    if result_idx + 1 < len(parts):
                        data_alarm = "Y"
                    break
                result_idx += 1
            
            current_row_data = {
                'test_name': test_name,
                'result': result,
                'data_alarm': data_alarm,
                'rerun': rerun,
                'has_rerun': has_plus
            }
            
            # 다음 줄에서 Unit, AU, R.P Lot 정보 추출
            i += 1
            if i < len(lines):
                next_line = lines[i].strip()
                next_parts = next_line.split()
                
                # 다음 줄의 첫 번째 단어를 Unit으로 사용
                if next_parts:
                    unit = next_parts[0]
                    
                    # E열(AU) 로직: 단위 다음 단어에 "-"가 없으면 같은 줄에서 "-" 포함된 단어 찾기
                    au = ""
                    
                    # 두번째 문장이 "NACL"인 경우 특수 처리
                    if len(next_parts) > 2 and next_parts[1] == "NACL":
                        # NACL인 경우: 세번째 단어 확인
                        if len(next_parts) > 2:
                            potential_au = next_parts[2]
                            if "-" in potential_au:
                                au = potential_au
                            else:
                                # 세번째 단어에 "-"가 없으면 같은 줄에서 "-" 포함된 단어 찾기
                                for word in next_parts:
                                    if "-" in word:
                                        au = word
                                        break
                        
                        rp_lot = ""
                        if len(next_parts) >= 6:  # 다섯번째 문장 (인덱스 4)
                            potential_rp_lot = next_parts[4]
                            # 숫자로만 구성되어 있거나 숫자가 포함된 문자열인 경우만 R.P Lot으로 인식
                            if potential_rp_lot.isdigit() or (potential_rp_lot.isalnum() and any(c.isdigit() for c in potential_rp_lot)):
                                rp_lot = potential_rp_lot
                    else:
                        # 일반적인 경우: 두번째 단어 확인
                        if len(next_parts) > 1:
                            potential_au = next_parts[1]
                            if "-" in potential_au:
                                au = potential_au
                            else:
                                # 두번째 단어에 "-"가 없으면 같은 줄에서 "-" 포함된 단어 찾기
                                for word in next_parts:
                                    if "-" in word:
                                        au = word
                                        break
                        
                        rp_lot = ""
                        if len(next_parts) >= 5 and len(next_parts) > 3:
                            potential_rp_lot = next_parts[3]
                            # 숫자로만 구성되어 있거나 숫자가 포함된 문자열인 경우만 R.P Lot으로 인식
                            if potential_rp_lot.isdigit() or (potential_rp_lot.isalnum() and any(c.isdigit() for c in potential_rp_lot)):
                                rp_lot = potential_rp_lot
                    
                    # COI인 경우 R/NR 값 추출
                    r_nr_value = ""
                    if unit == "COI":
                        # 다음 줄에서 Reac 또는 NonReac 찾기
                        if i + 1 < len(lines):
                            next_next_line = lines[i + 1].strip()
                            if "Reac" in next_next_line:
                                if "NonReac" in next_next_line:
                                    r_nr_value = "NonReac"
                                else:
                                    r_nr_value = "Reac"
                    
                    # 현재 행 데이터 완성
                    row_data = {
                        'sample_id': sample_id,
                        'test_name': current_row_data.get('test_name', ''),
                        'result': current_row_data.get('result', ''),
                        'unit': unit,
                        'au': au,
                        'rp_lot': rp_lot,
                        'data_alarm': current_row_data.get('data_alarm', 'N'),
                        'rerun': current_row_data.get('rerun', 'N'),
                        'date': date,
                        'has_rerun': current_row_data.get('has_rerun', False),
                        'r_nr': r_nr_value
                    }
                    extracted_data.append(row_data)
                    current_row_data = {}  # 다음 데이터를 위해 초기화
        
        i += 1
    
    return sample_id, date, extracted_data

def create_excel_file(pdf_filename, extracted_data, output_path, terminal_logs=None, pdf_lines=None):
    """
    추출된 데이터로 엑셀 파일을 생성하는 함수
    
    Args:
        pdf_filename (str): PDF 파일명 (시트명으로 사용)
        extracted_data (list): 추출된 데이터 리스트
        output_path (str): 출력 엑셀 파일 경로
        terminal_logs (list): 터미널 로그 리스트
        pdf_lines (list): PDF의 모든 줄 데이터 리스트
    """
    
    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    
    # 시트명 설정 (PDF 파일명에서 확장자 제거)
    sheet_name = os.path.splitext(pdf_filename)[0]
    ws.title = sheet_name
    
    # 헤더 설정
    headers = ['Sample ID', 'Test Name', 'Result', 'Unit', 'AU', 'R.P Lot', 'Data Alarm', 'Rerun', 'Date', 'R/NR']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # 데이터 입력
    for row_idx, data in enumerate(extracted_data, 2):
        # 안전한 데이터 처리
        ws.cell(row=row_idx, column=1, value=str(data.get('sample_id', '')) if data.get('sample_id') else '')      # A열: Sample ID
        ws.cell(row=row_idx, column=2, value=str(data.get('test_name', '')) if data.get('test_name') else '')   # B열: Test Name
        
        # Result 값을 숫자로 변환하여 입력
        try:
            # 안전한 문자열 처리
            result_str = str(data.get('result', '')) if data.get('result') is not None else ''
            if result_str and result_str.strip():
                result_value = float(result_str)
                # 유효숫자에 맞게 포맷팅
                if result_value == int(result_value):
                    # 정수인 경우 소수점 없이
                    result_value = int(result_value)
                else:
                    # 소수인 경우 적절한 자릿수로 반올림
                    if result_value >= 1:
                        # 1 이상인 경우 소수점 2자리까지
                        result_value = round(result_value, 2)
                    elif result_value >= 0.1:
                        # 0.1 이상인 경우 소수점 3자리까지
                        result_value = round(result_value, 3)
                    else:
                        # 0.1 미만인 경우 소수점 4자리까지
                        result_value = round(result_value, 4)
            else:
                result_value = ""  # 빈 값인 경우
        except (ValueError, TypeError, AttributeError):
            result_value = str(data.get('result', '')) if data.get('result') is not None else ''
        ws.cell(row=row_idx, column=3, value=result_value)        # C열: Result
        
        ws.cell(row=row_idx, column=4, value=str(data.get('unit', '')) if data.get('unit') else '')        # D열: Unit
        ws.cell(row=row_idx, column=5, value=str(data.get('au', '')) if data.get('au') else '')          # E열: AU
        ws.cell(row=row_idx, column=6, value=str(data.get('rp_lot', '')) if data.get('rp_lot') else '')  # F열: R.P Lot (없으면 공백)
        ws.cell(row=row_idx, column=7, value=str(data.get('data_alarm', 'N')) if data.get('data_alarm') else 'N')  # G열: Data Alarm
        ws.cell(row=row_idx, column=8, value=str(data.get('rerun', 'N')) if data.get('rerun') else 'N')       # H열: Rerun
        ws.cell(row=row_idx, column=9, value=str(data.get('date', '')) if data.get('date') else '')        # I열: Date
        ws.cell(row=row_idx, column=10, value=str(data.get('r_nr', '')) if data.get('r_nr') else '')        # J열: R/NR
        
        # Result 컬럼 스타일 적용
        result_cell = ws.cell(row=row_idx, column=3)
        result_cell.number_format = 'General'  # Result 컬럼을 일반형 서식으로 설정 (유효숫자만 표시)
        data_alarm_cell = ws.cell(row=row_idx, column=7)
        
        # Data Alarm이 Y인 경우 처리
        if data.get('data_alarm') == 'Y':
            data_alarm_cell.font = Font(color="FF0000", bold=True)  # Data Alarm 빨간색 굵게
            if not data.get('has_rerun', False):  # Rerun이 없는 경우만 Result를 빨간색으로
                result_cell.font = Font(color="FF0000", bold=True)
        
        # Rerun이 Y인 경우 Result를 연한 노란색 배경으로
        if data.get('rerun') == 'Y':
            result_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # 1행 전체에 필터 적용
    if len(extracted_data) > 0:
        # 헤더 행부터 마지막 데이터 행까지의 범위에 필터 적용
        last_row = len(extracted_data) + 1  # 헤더(1행) + 데이터 행 수
        last_col = len(headers)  # 컬럼 수
        ws.auto_filter.ref = f"A1:{chr(64 + last_col)}{last_row}"
    
    # 터미널 시트 추가 (PDF 줄별 내용)
    if pdf_lines:
        terminal_ws = wb.create_sheet(title="터미널 시트")
        
        # 헤더 설정
        terminal_ws.cell(row=1, column=1, value="페이지")
        terminal_ws.cell(row=1, column=2, value="줄 번호")
        terminal_ws.cell(row=1, column=3, value="내용")
        terminal_ws.cell(row=1, column=1).font = Font(bold=True)
        terminal_ws.cell(row=1, column=2).font = Font(bold=True)
        terminal_ws.cell(row=1, column=3).font = Font(bold=True)
        
        # PDF 줄별 데이터 입력
        row_idx = 2
        for page_data in pdf_lines:
            page_num = page_data.get('page', 1)
            lines = page_data.get('lines', [])
            
            for line_num, line_content in enumerate(lines, 1):
                try:
                    # 안전한 문자열 처리 (32000자 제한, 특수문자 제거)
                    safe_content = str(line_content)[:32000] if line_content else ""
                    safe_content = safe_content.replace('\x00', '').replace('\r', '').strip()
                    
                    terminal_ws.cell(row=row_idx, column=1, value=page_num)
                    terminal_ws.cell(row=row_idx, column=2, value=line_num)
                    terminal_ws.cell(row=row_idx, column=3, value=safe_content)
                    row_idx += 1
                except Exception as e:
                    # 오류 발생 시 안전한 처리
                    terminal_ws.cell(row=row_idx, column=1, value=page_num)
                    terminal_ws.cell(row=row_idx, column=2, value=line_num)
                    terminal_ws.cell(row=row_idx, column=3, value=f"[줄 처리 오류: {str(e)[:100]}]")
                    row_idx += 1
        
        # 컬럼 너비 조정
        terminal_ws.column_dimensions['A'].width = 10  # 페이지
        terminal_ws.column_dimensions['B'].width = 10  # 줄 번호
        terminal_ws.column_dimensions['C'].width = 100  # 내용
    
    # 터미널 로그 시트 추가 (기존 로직 유지)
    if terminal_logs:
        log_ws = wb.create_sheet(title="터미널 로그")
        
        # 헤더 설정
        log_ws.cell(row=1, column=1, value="터미널 로그")
        log_ws.cell(row=1, column=1).font = Font(bold=True)
        
        # 터미널 로그 데이터 입력
        for row_idx, log_line in enumerate(terminal_logs, 2):
            try:
                # 안전한 문자열 처리 (32000자 제한, 특수문자 제거)
                safe_log = str(log_line)[:32000] if log_line else ""
                safe_log = safe_log.replace('\x00', '').replace('\r', '').strip()
                log_ws.cell(row=row_idx, column=1, value=safe_log)
            except Exception as e:
                # 오류 발생 시 안전한 처리
                log_ws.cell(row=row_idx, column=1, value=f"[로그 처리 오류: {str(e)[:100]}]")
        
        # 컬럼 너비 조정
        log_ws.column_dimensions['A'].width = 100
    
    # 파일 저장
    wb.save(output_path)
    # 이 print는 create_excel_file 함수 내부이므로 여기서는 그대로 유지
    print(f"엑셀 파일이 저장되었습니다: {output_path}")

def select_save_location(pdf_filename):
    """
    GUI로 엑셀 파일 저장 위치를 선택하는 함수
    
    Args:
        pdf_filename (str): PDF 파일명
        
    Returns:
        str: 선택된 저장 경로, 취소시 None
    """
    if not TKINTER_AVAILABLE:
        return None
    
    # tkinter 윈도우 생성 (숨김)
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    
    # 기본 파일명 설정 (PDF 파일명에서 확장자 제거 후 _extracted.xlsx 추가)
    default_filename = os.path.splitext(pdf_filename)[0] + '_extracted.xlsx'
    
    # 마지막으로 사용한 디렉토리 불러오기
    initial_dir = load_last_directory()
    
    # 파일 저장 대화상자 열기
    save_path = filedialog.asksaveasfilename(
        title="엑셀 파일을 저장할 위치를 선택하세요",
        defaultextension=".xlsx",
        filetypes=[
            ("Excel 파일", "*.xlsx"),
            ("모든 파일", "*.*")
        ],
        initialfile=default_filename,
        initialdir=initial_dir
    )
    
    root.destroy()  # tkinter 윈도우 제거
    
    # 파일이 저장되었으면 해당 디렉토리를 저장
    if save_path:
        directory = os.path.dirname(save_path)
        save_last_directory(directory)
    
    return save_path if save_path else None

def open_excel_file(file_path):
    """
    엑셀 파일을 운영체제 기본 프로그램으로 열기
    
    Args:
        file_path (str): 열려는 엑셀 파일 경로
    """
    try:
        if platform.system() == "Windows":
            os.startfile(file_path)
        elif platform.system() == "Darwin":  # macOS
            subprocess.run(["open", file_path])
        else:  # Linux
            subprocess.run(["xdg-open", file_path])
        
        print(f"엑셀 파일이 열렸습니다: {file_path}")
    except Exception as e:
        print(f"엑셀 파일을 여는 중 오류가 발생했습니다: {str(e)}")
        print(f"수동으로 파일을 열어주세요: {file_path}")

def process_pdf_to_excel(pdf_path, progress_window=None):
    """
    PDF 파일을 읽어서 엑셀로 변환하는 메인 처리 함수
    
    Args:
        pdf_path (str): PDF 파일 경로
        progress_window (ProgressWindow): 프로그래스바 객체
    """
    
    # 터미널 로그 수집용 리스트
    terminal_logs = []
    # PDF 줄별 데이터 수집용 리스트
    pdf_lines = []
    
    def log_and_print(message):
        """터미널에 출력하면서 동시에 로그에 저장하는 함수"""
        print(message)
        terminal_logs.append(message)
    
    if not os.path.exists(pdf_path):
        log_and_print(f"오류: '{pdf_path}' 파일을 찾을 수 없습니다.")
        if progress_window:
            progress_window.close()
        return
    
    try:
        if progress_window:
            progress_window.update_progress(5, "Opening PDF file...")
        
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            if total_pages == 0:
                print("PDF에 페이지가 없습니다.")
                if progress_window:
                    progress_window.close()
                return
            
            if progress_window:
                progress_window.update_progress(10, f"Analyzing PDF pages... (Total {total_pages} pages)")
            
            log_and_print(f"PDF 총 페이지 수: {total_pages}")
            all_extracted_data = []
            sample_id = None
            date = None
            
            # 첫 번째 페이지 처리
            first_page = pdf.pages[0]
            text = first_page.extract_text()
            
            if not text:
                log_and_print("첫 번째 페이지에서 텍스트를 추출할 수 없습니다.")
                if progress_window:
                    progress_window.close()
                return
            
            lines = text.split('\n')
            
            # PDF 줄별 데이터 수집 (첫 번째 페이지)
            pdf_lines.append({
                'page': 1,
                'lines': [line.strip() for line in lines if line.strip()]
            })
            
            if progress_window:
                progress_window.update_progress(20, "Extracting data from first page...")
            
            # 디버깅용: 줄 번호와 내용 출력
            log_and_print("=" * 50)
            log_and_print("첫 번째 페이지 내용:")
            log_and_print("=" * 50)
            for i, line in enumerate(lines, 1):
                if line.strip():
                    log_and_print(f"줄 {i:3d}: {line}")
            log_and_print("=" * 50)
            
            # 첫 번째 페이지 데이터 추출
            sample_id, date, first_page_data = extract_data_from_first_page(lines)
            all_extracted_data.extend(first_page_data)
            
            if progress_window:
                progress_window.update_progress(30, f"First page completed ({len(first_page_data)} data items)")
            
            log_and_print(f"\n첫 번째 페이지에서 추출된 데이터: {len(first_page_data)}개")
            
            # 두 번째 페이지부터 처리
            for page_num in range(1, total_pages):
                # 진행률 계산 (30%부터 60%까지)
                if progress_window and total_pages > 1:
                    progress = 30 + int((page_num / (total_pages - 1)) * 30)
                    progress_window.update_progress(progress, f"Processing page {page_num + 1}/{total_pages}...")
                
                page = pdf.pages[page_num]
                text = page.extract_text()
                
                if not text:
                    log_and_print(f"페이지 {page_num + 1}에서 텍스트를 추출할 수 없습니다.")
                    continue
                
                lines = text.split('\n')
                
                # PDF 줄별 데이터 수집 (다른 페이지들)
                pdf_lines.append({
                    'page': page_num + 1,
                    'lines': [line.strip() for line in lines if line.strip()]
                })
                
                # 디버깅용: 줄 번호와 내용 출력
                log_and_print(f"\n[ 페이지 {page_num + 1} ]")
                log_and_print("-" * 30)
                for i, line in enumerate(lines, 1):
                    if line.strip():
                        log_and_print(f"줄 {i:3d}: {line}")
                
                # 두 번째 페이지부터의 데이터 추출
                page_sample_id, page_date, page_data = extract_data_from_other_pages(lines)
                all_extracted_data.extend(page_data)
                
                log_and_print(f"페이지 {page_num + 1}에서 추출된 데이터: {len(page_data)}개")
                log_and_print(f"  - Sample ID: {page_sample_id}, Date: {page_date}")
            
            if not all_extracted_data:
                log_and_print("추출할 데이터가 없습니다.")
                if progress_window:
                    progress_window.close()
                return
            
            if progress_window:
                progress_window.update_progress(60, "Organizing data...")
            
            log_and_print(f"\n전체 추출된 데이터:")
            log_and_print(f"Sample ID: {sample_id}")
            log_and_print(f"Date: {date}")
            log_and_print(f"총 데이터 개수: {len(all_extracted_data)}")
            
            # 데이터 출력 (디버깅용)
            for i, data in enumerate(all_extracted_data, 1):
                log_and_print(f"  {i:2d}. Sample ID: {data.get('sample_id', '')}, Test Name: {data.get('test_name', '')}, Result: {data.get('result', '')}, Unit: {data.get('unit', '')}, AU: {data.get('au', '')}")
            
            if progress_window:
                progress_window.update_progress(70, "Selecting output location...")
            
            # 저장 위치 선택
            pdf_filename = os.path.basename(pdf_path)
            output_path = select_save_location(pdf_filename)
            
            if not output_path:
                log_and_print("저장이 취소되었습니다.")
                if progress_window:
                    progress_window.close()
                return
            
            if progress_window:
                progress_window.update_progress(80, "Creating Excel file...")
            
            # 엑셀 파일 생성 (PDF 줄별 데이터 포함)
            create_excel_file(pdf_filename, all_extracted_data, output_path, terminal_logs, pdf_lines)
            
            if progress_window:
                progress_window.update_progress(100, "Completed!")
            
            log_and_print(f"\n변환 완료!")
            log_and_print(f"출력 파일: {output_path}")
            
    except Exception as e:
        log_and_print(f"PDF 처리 중 오류 발생: {e}")
        import traceback
        log_and_print(f"상세 오류: {traceback.format_exc()}")
    finally:
        if progress_window:
            progress_window.close()

def select_pdf_file():
    """
    GUI로 PDF 파일을 선택하는 함수
    
    Returns:
        str: 선택된 PDF 파일의 경로, 취소시 None
    """
    if not TKINTER_AVAILABLE:
        return None
    
    # tkinter 윈도우 생성 (숨김)
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    
    # 마지막으로 사용한 디렉토리 불러오기
    initial_dir = load_last_directory()
    
    # 파일 선택 대화상자 열기
    pdf_path = filedialog.askopenfilename(
        title="PDF 파일을 선택하세요",
        filetypes=[
            ("PDF 파일", "*.pdf"),
            ("모든 파일", "*.*")
        ],
        initialdir=initial_dir  # 마지막 사용 디렉토리에서 시작
    )
    
    root.destroy()  # tkinter 윈도우 제거
    
    # 파일이 선택되었으면 해당 디렉토리를 저장
    if pdf_path:
        directory = os.path.dirname(pdf_path)
        save_last_directory(directory)
    
    return pdf_path if pdf_path else None

def run(pdf_path:str) -> str:
    """
    Entrypoint: converts PDF to Excel and returns output path
    Streamlit 환경에서 호출될 때는 파일 저장 대화상자를 표시하지 않고 임시 파일에 저장합니다.
    
    Args:
        pdf_path (str): PDF 파일 경로
        
    Returns:
        str: 생성된 Excel 파일 경로
    """
    # Streamlit 환경에서 실행 중인지 확인
    is_streamlit = 'streamlit' in sys.modules
    
    # 터미널 로그를 저장할 리스트
    terminal_logs = []
    # PDF 줄별 데이터 수집용 리스트
    pdf_lines = []
    
    def log_and_print(msg):
        terminal_logs.append(msg)
        print(msg)
    
    # 입력 파일 체크
    if not os.path.exists(pdf_path):
        log_and_print(f"오류: 파일을 찾을 수 없습니다: {pdf_path}")
        return None

    try:
        # PDF 열기
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            if total_pages == 0:
                log_and_print("PDF에 페이지가 없습니다.")
                return None

            # 첫 페이지 추출
            lines = pdf.pages[0].extract_text().split('\n')
            
            # PDF 줄별 데이터 수집 (첫 번째 페이지)
            pdf_lines.append({
                'page': 1,
                'lines': [line.strip() for line in lines if line.strip()]
            })
            
            sample_id, date, extracted = extract_data_from_first_page(lines)

            # 이후 페이지 추출
            for i, page in enumerate(pdf.pages[1:], start=1):
                lines = page.extract_text().split('\n')
                
                # PDF 줄별 데이터 수집 (다른 페이지들)
                pdf_lines.append({
                    'page': i + 1,
                    'lines': [line.strip() for line in lines if line.strip()]
                })
                
                _, _, data = extract_data_from_other_pages(lines)
                extracted.extend(data)

        if not extracted:
            log_and_print("추출된 데이터가 없습니다.")
            return None

        # Streamlit 환경에서는 임시 파일에 저장
        if is_streamlit:
            import tempfile
            pdf_filename = os.path.basename(pdf_path)
            base_name = os.path.splitext(pdf_filename)[0]
            # 임시 파일 생성
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                output_path = tmp.name
            
            # 파일 이름을 PDF 파일명과 동일하게 설정 (확장자만 .xlsx로 변경)
            new_output_path = os.path.join(os.path.dirname(output_path), f"{base_name}.xlsx")
            if os.path.exists(new_output_path):
                try:
                    os.remove(new_output_path)  # 기존 파일이 있으면 삭제
                except:
                    pass
            try:
                os.rename(output_path, new_output_path)
                output_path = new_output_path
            except:
                # 이름 변경 실패 시 원래 임시 파일 경로 사용
                pass
        else:
            # 일반 환경에서는 사용자에게 저장 위치 선택 요청
            pdf_filename = os.path.basename(pdf_path)
            output_path = select_save_location(pdf_filename)
            if not output_path:
                return None

        # 엑셀 생성 (PDF 줄별 데이터 포함)
        create_excel_file(os.path.basename(pdf_path), extracted, output_path, terminal_logs, pdf_lines)
        return output_path

    except Exception as e:
        log_and_print(f"PDF 처리 중 오류 발생: {e}")
        return None

def main():
    """
    메인 함수: GUI로 PDF 파일을 선택받아 엑셀로 변환합니다.
    """
    
    if len(sys.argv) > 1:
        # 명령행 인수로 파일 경로가 제공된 경우
        pdf_path = sys.argv[1]
        print(f"명령행에서 제공된 파일: {pdf_path}")
    else:
        # GUI로 파일 선택
        print("PDF 파일 선택 창을 열고 있습니다...")
        pdf_path = select_pdf_file()
        
        if not pdf_path:
            print("파일 선택이 취소되었습니다.")
            return
            
        print(f"선택된 파일: {pdf_path}")
    
    # 프로그래스바 생성 및 표시
    progress_window = ProgressWindow()
    progress_window.show()
    
    # PDF를 엑셀로 변환
    process_pdf_to_excel(pdf_path, progress_window)

if __name__ == "__main__":
    main()
