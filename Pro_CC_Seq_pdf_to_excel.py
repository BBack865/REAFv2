import pdfplumber
import sys
import os
# Streamlit 환경에서는 tkinter를 사용하지 않음
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    TKINTER_AVAILABLE = True
except ImportError:
    TKINTER_AVAILABLE = False
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import re
import subprocess
import platform
import threading
import time
import json

def get_config_file_path():
    """
    설정 파일 경로를 반환하는 함수
    
    Returns:
        str: 설정 파일의 전체 경로
    """
    # 현재 스크립트와 같은 디렉토리에 config.json 파일 생성
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(script_dir, "pdf_converter_config.json")

def load_last_directory():
    """
    마지막으로 사용한 디렉토리를 불러오는 함수
    
    Returns:
        str: 마지막 사용 디렉토리 경로, 없으면 현재 디렉토리
    """
    config_file = get_config_file_path()
    try:
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                last_dir = config.get('last_pdf_directory', os.getcwd())
                # 디렉토리가 실제로 존재하는지 확인
                if os.path.exists(last_dir):
                    return last_dir
    except (json.JSONDecodeError, FileNotFoundError, PermissionError):
        pass
    
    return os.getcwd()

def save_last_directory(directory_path):
    """
    마지막으로 사용한 디렉토리를 저장하는 함수
    
    Args:
        directory_path (str): 저장할 디렉토리 경로
    """
    config_file = get_config_file_path()
    try:
        # 기존 설정이 있으면 불러오기
        config = {}
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            except (json.JSONDecodeError, FileNotFoundError):
                config = {}
        
        # 마지막 디렉토리 업데이트
        config['last_pdf_directory'] = directory_path
        
        # 설정 파일에 저장
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except (PermissionError, OSError):
        # 저장 실패시 무시 (치명적이지 않음)
        pass

class ProgressWindow:
    """
    프로그래스바를 표시하는 GUI 클래스
    """
    def __init__(self):
        if not TKINTER_AVAILABLE:
            return
        
        self.root = tk.Tk()
        self.root.title("PDF 처리 중...")
        
        # 창 크기 설정
        window_width = 400
        window_height = 150
        
        # 화면 크기 가져오기
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # 화면 중앙 위치 계산
        center_x = int(screen_width/2 - window_width/2)
        center_y = int(screen_height/2 - window_height/2)
        
        # 창 크기와 위치 설정
        self.root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
        self.root.resizable(False, False)
        self.root.attributes('-topmost', True)  # 항상 위에 표시
        
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 제목 라벨
        self.title_label = ttk.Label(main_frame, text="Processing PDF file...", 
                                    font=("Arial", 12, "bold"))
        self.title_label.pack(pady=(0, 20))
        
        # 상태 라벨
        self.status_label = ttk.Label(main_frame, text="Starting...", 
                                     font=("Arial", 10),
                                     width=50,
                                     anchor="center")
        self.status_label.pack(pady=(0, 10))
        
        # 프로그래스바
        self.progress = ttk.Progressbar(main_frame, length=300, mode='determinate')
        self.progress.pack(pady=(0, 10))
        
        # 퍼센트 라벨
        self.percent_label = ttk.Label(main_frame, text="0%", 
                                      font=("Arial", 10),
                                      width=10,
                                      anchor="center")
        self.percent_label.pack()
        
        # 프로그래스바 초기화
        self.progress['maximum'] = 100
        self.progress['value'] = 0
        
    def update_progress(self, value, status_text=""):
        """
        프로그래스바 업데이트
        
        Args:
            value (int): 프로그래스 값 (0-100)
            status_text (str): 상태 텍스트
        """
        if not TKINTER_AVAILABLE:
            return
        
        self.progress['value'] = value
        self.percent_label.config(text=f"{value}%")
        if status_text:
            # 이전 텍스트를 완전히 지우고 새 텍스트로 교체
            self.status_label.config(text="")
            self.root.update_idletasks()
            self.status_label.config(text=status_text)
        self.root.update_idletasks()
        
    def close(self):
        """프로그래스바 창 닫기"""
        if not TKINTER_AVAILABLE:
            return
        
        self.root.destroy()
        
    def show(self):
        """프로그래스바 창 표시"""
        if not TKINTER_AVAILABLE:
            return
        
        self.root.update()

def extract_data_from_first_page(lines):
    """
    첫 번째 페이지의 특정 줄에서 데이터를 추출하는 함수
    8번째 줄에서 Seq No.와 Date 추출, 13~30번째 줄에서 데이터 처리
    
    Args:
        lines (list): 페이지의 모든 줄들
        
    Returns:
        tuple: (base_seq_no, date, extracted_data, test_counter)
    """
    base_seq_no = None
    date = None
    extracted_data = []
    current_row_data = {}
    test_counter = 0  # 테스트 순서 카운터 추가
    
    # 8번째 줄에서 Seq No.와 Date 추출 (인덱스 7)
    if len(lines) > 7:
        line_8 = lines[7].strip()
        if "Ser/PI" in line_8 or "SerumPlasma" in line_8:
            parts = line_8.split()
            if len(parts) >= 2:
                base_seq_no = parts[1]  # 두 번째 문단
                # YYYY/MM/DD 형태의 날짜 찾기
                for part in parts:
                    if re.match(r'\d{4}/\d{2}/\d{2}', part):
                        date = part
                        break
    
    # 13번째 줄부터 30번째 줄까지 처리 (인덱스 12부터 29까지)
    start_line = 12  # 13번째 줄 (0-based index)
    end_line = min(30, len(lines))  # 30번째 줄까지 또는 페이지 끝까지
    
    i = start_line
    while i < end_line:
        if i >= len(lines):
            break
            
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        
        parts = line.split()
        if not parts:
            i += 1
            continue
        
        # "R2" 또는 "R3"로 시작하는 줄은 생략
        if parts[0] in ["R2", "R3"]:
            i += 1
            continue
        
        # Test Name과 Result 패턴 처리
        if re.match(r'^[\+]?[A-Z][A-Z0-9\-]*\s+[\d\.]+', line) or line.startswith('+') or line.startswith('ISE'):
            # "+" 존재 여부 확인
            has_plus = line.startswith('+')
            
            # ISE 특별 처리
            if parts[0] == "ISE" or (has_plus and len(parts) > 1 and parts[1] == "ISE"):
                if has_plus:
                    # "+ ISE K 4.5" 형태
                    if len(parts) >= 4:
                        test_name = f"{parts[1]} {parts[2]}"  # "ISE K"
                        result = parts[3]  # "4.5"
                        # Result 뒤에 추가 단어가 있으면 Data Alarm Y
                        data_alarm = "Y" if len(parts) > 4 else "N"
                        rerun = "Y"
                        
                        current_row_data = {
                            'test_name': test_name,
                            'result': result,
                            'data_alarm': data_alarm,
                            'rerun': rerun,
                            'has_rerun': True
                        }
                else:
                    # "ISE K 4.5" 형태
                    if len(parts) >= 3:
                        test_name = f"{parts[0]} {parts[1]}"  # "ISE K"
                        result = parts[2]  # "4.5"
                        # Result 뒤에 추가 단어가 있으면 Data Alarm Y
                        data_alarm = "Y" if len(parts) > 3 else "N"
                        rerun = "N"
                        
                        current_row_data = {
                            'test_name': test_name,
                            'result': result,
                            'data_alarm': data_alarm,
                            'rerun': rerun,
                            'has_rerun': False
                        }
            elif has_plus:
                # "+ BILD2-D 0.627 > Test" 형태
                if len(parts) >= 3:
                    test_name = parts[1]
                    result = parts[2]
                    # Result 뒤에 추가 단어가 있으면 Data Alarm Y
                    data_alarm = "Y" if len(parts) > 3 else "N"
                    rerun = "Y"
                    
                    current_row_data = {
                        'test_name': test_name,
                        'result': result,
                        'data_alarm': data_alarm,
                        'rerun': rerun,
                        'has_rerun': True
                    }
            else:
                # "BILD2-D 0.627" 또는 "BILD2-D 0.627 > Test" 형태
                if len(parts) >= 2:
                    test_name = parts[0]
                    result = parts[1]
                    # Result 뒤에 추가 단어가 있으면 Data Alarm Y
                    data_alarm = "Y" if len(parts) > 2 else "N"
                    rerun = "N"
                    
                    current_row_data = {
                        'test_name': test_name,
                        'result': result,
                        'data_alarm': data_alarm,
                        'rerun': rerun,
                        'has_rerun': False
                    }
            
            # 다음 줄에서 Unit, AU, R.P Lot 정보 추출
            i += 1
            if i < len(lines):
                next_line = lines[i].strip()
                next_parts = next_line.split()
                
                if next_parts and any(unit in next_line for unit in ['mg/dL', 'g/dL', 'mmol/L', 'U/L', '%']):
                    if len(next_parts) >= 2:
                        unit = next_parts[0]
                        
                        # 두번째 문장이 "NACL"인 경우 특수 처리
                        if len(next_parts) > 2 and next_parts[1] == "NACL":
                            # NACL인 경우: 세번째 문장이 AU, 다섯번째 문장이 R.P Lot
                            au = next_parts[2] if len(next_parts) > 2 else ""
                            rp_lot = ""
                            if len(next_parts) >= 6:  # 다섯번째 문장 (인덱스 4)
                                potential_rp_lot = next_parts[4]
                                # 숫자로만 구성되어 있거나 숫자가 포함된 문자열인 경우만 R.P Lot으로 인식
                                if potential_rp_lot.isdigit() or (potential_rp_lot.isalnum() and any(c.isdigit() for c in potential_rp_lot)):
                                    rp_lot = potential_rp_lot
                        else:
                            # 일반적인 경우: 두번째 문장이 AU, 네번째 문장이 R.P Lot
                            au = next_parts[1]
                            rp_lot = ""
                            if len(next_parts) >= 5 and len(next_parts) > 3:
                                potential_rp_lot = next_parts[3]
                                # 숫자로만 구성되어 있거나 숫자가 포함된 문자열인 경우만 R.P Lot으로 인식
                                if potential_rp_lot.isdigit() or (potential_rp_lot.isalnum() and any(c.isdigit() for c in potential_rp_lot)):
                                    rp_lot = potential_rp_lot
                        
                        # 개별 순차 번호 생성
                        test_counter += 1
                        if base_seq_no:
                            # 기본 seq_no에서 숫자 부분 추출하여 증가
                            try:
                                base_num = int(base_seq_no)
                                individual_seq_no = f"{base_num + test_counter - 1:06d}"
                            except ValueError:
                                # 숫자가 아닌 경우 그대로 사용하고 카운터 추가
                                individual_seq_no = f"{base_seq_no}-{test_counter}"
                        else:
                            individual_seq_no = f"{test_counter:06d}"
                        
                        # 현재 행 데이터 완성
                        row_data = {
                            'seq_no': individual_seq_no,  # 개별 순차 번호 사용
                            'test_name': current_row_data.get('test_name', ''),
                            'result': current_row_data.get('result', ''),
                            'unit': unit,
                            'au': au,
                            'rp_lot': rp_lot,
                            'data_alarm': current_row_data.get('data_alarm', 'N'),
                            'rerun': current_row_data.get('rerun', 'N'),
                            'date': date,
                            'has_rerun': current_row_data.get('has_rerun', False)
                        }
                        extracted_data.append(row_data)
                        current_row_data = {}  # 다음 데이터를 위해 초기화
        
        i += 1
    
    return base_seq_no, date, extracted_data, test_counter

def extract_data_from_other_pages(lines, global_test_counter=0):
    """
    두 번째 페이지부터의 특정 줄에서 데이터를 추출하는 함수
    5번째 줄에서 Seq No.와 Date 추출, 10번째 줄부터 30번째 줄까지 데이터 처리
    
    Args:
        lines (list): 페이지의 모든 줄들
        global_test_counter (int): 전역 테스트 카운터 (페이지 간 연속성 유지)
        
    Returns:
        tuple: (base_seq_no, date, extracted_data, test_counter)
    """
    extracted_data = []
    current_row_data = {}
    base_seq_no = None
    date = None
    test_counter = global_test_counter  # 전역 카운터에서 시작
    
    # 5번째 줄에서 Seq No.와 Date 추출 (인덱스 4)
    if len(lines) > 4:
        line_5 = lines[4].strip()
        if "Ser/PI" in line_5:
            parts = line_5.split()
            if len(parts) >= 2:
                base_seq_no = parts[1]  # 두 번째 문단
                # YYYY/MM/DD 형태의 날짜 찾기
                for part in parts:
                    if re.match(r'\d{4}/\d{2}/\d{2}', part):
                        date = part
                        break
    
    # 10번째 줄부터 30번째 줄까지 처리 (인덱스 9부터 29까지)
    start_line = 9  # 10번째 줄 (0-based index)
    end_line = min(30, len(lines))  # 30번째 줄까지 또는 페이지 끝까지
    
    i = start_line
    while i < end_line:
        if i >= len(lines):
            break
            
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        
        parts = line.split()
        if not parts:
            i += 1
            continue
            
        # "R2" 또는 "R3"로 시작하는 줄은 생략
        if parts[0] in ["R2", "R3"]:
            i += 1
            continue
        
        # Test Name과 Result 패턴 처리
        if re.match(r'^[\+]?[A-Z][A-Z0-9\-]*\s+[\d\.]+', line) or line.startswith('+') or line.startswith('ISE'):
            # "+" 존재 여부 확인
            has_plus = line.startswith('+')
            
            # ISE 특별 처리
            if parts[0] == "ISE" or (has_plus and len(parts) > 1 and parts[1] == "ISE"):
                if has_plus:
                    # "+ ISE K 4.5" 형태
                    if len(parts) >= 4:
                        test_name = f"{parts[1]} {parts[2]}"  # "ISE K"
                        result = parts[3]  # "4.5"
                        # Result 뒤에 추가 단어가 있으면 Data Alarm Y
                        data_alarm = "Y" if len(parts) > 4 else "N"
                        rerun = "Y"
                        
                        current_row_data = {
                            'test_name': test_name,
                            'result': result,
                            'data_alarm': data_alarm,
                            'rerun': rerun,
                            'has_rerun': True
                        }
                else:
                    # "ISE K 4.5" 형태
                    if len(parts) >= 3:
                        test_name = f"{parts[0]} {parts[1]}"  # "ISE K"
                        result = parts[2]  # "4.5"
                        # Result 뒤에 추가 단어가 있으면 Data Alarm Y
                        data_alarm = "Y" if len(parts) > 3 else "N"
                        rerun = "N"
                        
                        current_row_data = {
                            'test_name': test_name,
                            'result': result,
                            'data_alarm': data_alarm,
                            'rerun': rerun,
                            'has_rerun': False
                        }
            elif has_plus:
                # "+ CHOL2-I 178 > Test" 형태
                if len(parts) >= 4:
                    test_name = parts[1]
                    result = parts[2]
                    # Result 뒤에 추가 단어가 있으면 Data Alarm Y
                    data_alarm = "Y" if len(parts) > 4 else "N"
                    rerun = "Y"
                    
                    current_row_data = {
                        'test_name': test_name,
                        'result': result,
                        'data_alarm': data_alarm,
                        'rerun': rerun,
                        'has_rerun': True
                    }
            else:
                # "CHOL2-I 178" 또는 "CHOL2-I 178 > Test" 형태
                if len(parts) >= 2:
                    test_name = parts[0]
                    result = parts[1]
                    # Result 뒤에 추가 단어가 있으면 Data Alarm Y
                    data_alarm = "Y" if len(parts) > 2 else "N"
                    rerun = "N"
                    
                    current_row_data = {
                        'test_name': test_name,
                        'result': result,
                        'data_alarm': data_alarm,
                        'rerun': rerun,
                        'has_rerun': False
                    }
            
            # 다음 줄에서 Unit, AU, R.P Lot 정보 추출
            i += 1
            if i < len(lines):
                next_line = lines[i].strip()
                next_parts = next_line.split()
                
                if next_parts and any(unit in next_line for unit in ['mg/dL', 'g/dL', 'mmol/L', 'U/L', '%']):
                    if len(next_parts) >= 2:
                        unit = next_parts[0]
                        
                        # 두번째 문장이 "NACL"인 경우 특수 처리
                        if len(next_parts) > 2 and next_parts[1] == "NACL":
                            # NACL인 경우: 세번째 문장이 AU, 다섯번째 문장이 R.P Lot
                            au = next_parts[2] if len(next_parts) > 2 else ""
                            rp_lot = ""
                            if len(next_parts) >= 6:  # 다섯번째 문장 (인덱스 4)
                                potential_rp_lot = next_parts[4]
                                # 숫자로만 구성되어 있거나 숫자가 포함된 문자열인 경우만 R.P Lot으로 인식
                                if potential_rp_lot.isdigit() or (potential_rp_lot.isalnum() and any(c.isdigit() for c in potential_rp_lot)):
                                    rp_lot = potential_rp_lot
                        else:
                            # 일반적인 경우: 두번째 문장이 AU, 네번째 문장이 R.P Lot
                            au = next_parts[1]
                            rp_lot = ""
                            if len(next_parts) >= 5 and len(next_parts) > 3:
                                potential_rp_lot = next_parts[3]
                                # 숫자로만 구성되어 있거나 숫자가 포함된 문자열인 경우만 R.P Lot으로 인식
                                if potential_rp_lot.isdigit() or (potential_rp_lot.isalnum() and any(c.isdigit() for c in potential_rp_lot)):
                                    rp_lot = potential_rp_lot
                        
                        # 개별 순차 번호 생성
                        test_counter += 1
                        if base_seq_no:
                            # 기본 seq_no에서 숫자 부분 추출하여 증가
                            try:
                                base_num = int(base_seq_no)
                                individual_seq_no = f"{base_num + test_counter - 1:06d}"
                            except ValueError:
                                # 숫자가 아닌 경우 그대로 사용하고 카운터 추가
                                individual_seq_no = f"{base_seq_no}-{test_counter}"
                        else:
                            individual_seq_no = f"{test_counter:06d}"
                        
                        # 현재 행 데이터 완성
                        row_data = {
                            'seq_no': individual_seq_no,  # 개별 순차 번호 사용
                            'test_name': current_row_data.get('test_name', ''),
                            'result': current_row_data.get('result', ''),
                            'unit': unit,
                            'au': au,
                            'rp_lot': rp_lot,
                            'data_alarm': current_row_data.get('data_alarm', 'N'),
                            'rerun': current_row_data.get('rerun', 'N'),
                            'date': date,
                            'has_rerun': current_row_data.get('has_rerun', False)
                        }
                        extracted_data.append(row_data)
                        current_row_data = {}  # 다음 데이터를 위해 초기화
        
        i += 1
    
    return base_seq_no, date, extracted_data, test_counter

def create_excel_file(pdf_filename, extracted_data, output_path, terminal_logs=None, pdf_lines=None):
    """
    추출된 데이터로 엑셀 파일을 생성하는 함수
    
    Args:
        pdf_filename (str): PDF 파일명 (시트명으로 사용)
        extracted_data (list): 추출된 데이터 리스트
        output_path (str): 출력 엑셀 파일 경로
        terminal_logs (list): 터미널 로그 리스트
        pdf_lines (list): PDF의 모든 줄 데이터 리스트
    """
    
    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    
    # 시트명 설정 (PDF 파일명에서 확장자 제거)
    sheet_name = os.path.splitext(pdf_filename)[0]
    ws.title = sheet_name
    
    # 헤더 설정
    headers = ['Seq No.', 'Test Name', 'Result', 'Unit', 'AU', 'R.P Lot', 'Data Alarm', 'Rerun', 'Date']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # 데이터 입력
    for row_idx, data in enumerate(extracted_data, 2):
        # 안전한 데이터 처리
        ws.cell(row=row_idx, column=1, value=str(data.get('seq_no', '')) if data.get('seq_no') else '')      # A열: Seq No.
        ws.cell(row=row_idx, column=2, value=str(data.get('test_name', '')) if data.get('test_name') else '')   # B열: Test Name
        
        # Result 값을 숫자로 변환하여 입력
        try:
            # 안전한 문자열 처리
            result_str = str(data.get('result', '')) if data.get('result') is not None else ''
            if result_str and result_str.strip():
                result_value = float(result_str)
                # 유효숫자에 맞게 포맷팅
                if result_value == int(result_value):
                    # 정수인 경우 소수점 없이
                    result_value = int(result_value)
                else:
                    # 소수인 경우 적절한 자릿수로 반올림
                    if result_value >= 1:
                        # 1 이상인 경우 소수점 2자리까지
                        result_value = round(result_value, 2)
                    elif result_value >= 0.1:
                        # 0.1 이상인 경우 소수점 3자리까지
                        result_value = round(result_value, 3)
                    else:
                        # 0.1 미만인 경우 소수점 4자리까지
                        result_value = round(result_value, 4)
            else:
                result_value = ""  # 빈 값인 경우
        except (ValueError, TypeError, AttributeError):
            result_value = str(data.get('result', '')) if data.get('result') is not None else ''
        ws.cell(row=row_idx, column=3, value=result_value)        # C열: Result
        
        ws.cell(row=row_idx, column=4, value=str(data.get('unit', '')) if data.get('unit') else '')        # D열: Unit
        ws.cell(row=row_idx, column=5, value=str(data.get('au', '')) if data.get('au') else '')          # E열: AU
        ws.cell(row=row_idx, column=6, value=str(data.get('rp_lot', '')) if data.get('rp_lot') else '')  # F열: R.P Lot (없으면 공백)
        ws.cell(row=row_idx, column=7, value=str(data.get('data_alarm', 'N')) if data.get('data_alarm') else 'N')  # G열: Data Alarm
        ws.cell(row=row_idx, column=8, value=str(data.get('rerun', 'N')) if data.get('rerun') else 'N')       # H열: Rerun
        ws.cell(row=row_idx, column=9, value=str(data.get('date', '')) if data.get('date') else '')        # I열: Date
        
        # Result 컬럼 스타일 적용
        result_cell = ws.cell(row=row_idx, column=3)
        result_cell.number_format = 'General'  # Result 컬럼을 일반형 서식으로 설정 (유효숫자만 표시)
        data_alarm_cell = ws.cell(row=row_idx, column=7)
        
        # Data Alarm이 Y인 경우 처리
        if data.get('data_alarm') == 'Y':
            data_alarm_cell.font = Font(color="FF0000", bold=True)  # Data Alarm 빨간색 굵게
            if not data.get('has_rerun', False):  # Rerun이 없는 경우만 Result를 빨간색으로
                result_cell.font = Font(color="FF0000", bold=True)
        
        # Rerun이 Y인 경우 Result를 연한 노란색 배경으로
        if data.get('rerun') == 'Y':
            result_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # 1행 전체에 필터 적용
    if len(extracted_data) > 0:
        # 헤더 행부터 마지막 데이터 행까지의 범위에 필터 적용
        last_row = len(extracted_data) + 1  # 헤더(1행) + 데이터 행 수
        last_col = len(headers)  # 컬럼 수
        ws.auto_filter.ref = f"A1:{chr(64 + last_col)}{last_row}"
    
    # 터미널 시트 추가 (PDF 줄별 내용)
    if pdf_lines:
        terminal_ws = wb.create_sheet(title="터미널 시트")
        
        # 헤더 설정
        terminal_ws.cell(row=1, column=1, value="페이지")
        terminal_ws.cell(row=1, column=2, value="줄 번호")
        terminal_ws.cell(row=1, column=3, value="내용")
        terminal_ws.cell(row=1, column=1).font = Font(bold=True)
        terminal_ws.cell(row=1, column=2).font = Font(bold=True)
        terminal_ws.cell(row=1, column=3).font = Font(bold=True)
        
        # PDF 줄별 데이터 입력
        row_idx = 2
        for page_data in pdf_lines:
            page_num = page_data.get('page', 1)
            lines = page_data.get('lines', [])
            
            for line_num, line_content in enumerate(lines, 1):
                try:
                    # 안전한 문자열 처리 (32000자 제한, 특수문자 제거)
                    safe_content = str(line_content)[:32000] if line_content else ""
                    safe_content = safe_content.replace('\x00', '').replace('\r', '').strip()
                    
                    terminal_ws.cell(row=row_idx, column=1, value=page_num)
                    terminal_ws.cell(row=row_idx, column=2, value=line_num)
                    terminal_ws.cell(row=row_idx, column=3, value=safe_content)
                    row_idx += 1
                except Exception as e:
                    # 오류 발생 시 안전한 처리
                    terminal_ws.cell(row=row_idx, column=1, value=page_num)
                    terminal_ws.cell(row=row_idx, column=2, value=line_num)
                    terminal_ws.cell(row=row_idx, column=3, value=f"[줄 처리 오류: {str(e)[:100]}]")
                    row_idx += 1
        
        # 컬럼 너비 조정
        terminal_ws.column_dimensions['A'].width = 10  # 페이지
        terminal_ws.column_dimensions['B'].width = 10  # 줄 번호
        terminal_ws.column_dimensions['C'].width = 100  # 내용
    
    # 터미널 로그 시트 추가 (기존 로직 유지)
    if terminal_logs:
        log_ws = wb.create_sheet(title="터미널 로그")
        
        # 헤더 설정
        log_ws.cell(row=1, column=1, value="터미널 로그")
        log_ws.cell(row=1, column=1).font = Font(bold=True)
        
        # 터미널 로그 데이터 입력
        for row_idx, log_line in enumerate(terminal_logs, 2):
            try:
                # 안전한 문자열 처리 (32000자 제한, 특수문자 제거)
                safe_log = str(log_line)[:32000] if log_line else ""
                safe_log = safe_log.replace('\x00', '').replace('\r', '').strip()
                log_ws.cell(row=row_idx, column=1, value=safe_log)
            except Exception as e:
                # 오류 발생 시 안전한 처리
                log_ws.cell(row=row_idx, column=1, value=f"[로그 처리 오류: {str(e)[:100]}]")
        
        # 컬럼 너비 조정
        log_ws.column_dimensions['A'].width = 100
    
    # 파일 저장
    wb.save(output_path)
    print(f"엑셀 파일이 저장되었습니다: {output_path}")

def select_save_location(pdf_filename):
    """
    GUI로 엑셀 파일 저장 위치를 선택하는 함수
    
    Args:
        pdf_filename (str): PDF 파일명
        
    Returns:
        str: 선택된 저장 경로, 취소시 None
    """
    if not TKINTER_AVAILABLE:
        return None
    
    # tkinter 윈도우 생성 (숨김)
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    
    # 기본 파일명 설정 (PDF 파일명에서 확장자 제거 후 _extracted.xlsx 추가)
    default_filename = os.path.splitext(pdf_filename)[0] + '_extracted.xlsx'
    
    # 마지막으로 사용한 디렉토리 불러오기
    initial_dir = load_last_directory()
    
    # 파일 저장 대화상자 열기
    save_path = filedialog.asksaveasfilename(
        title="엑셀 파일을 저장할 위치를 선택하세요",
        defaultextension=".xlsx",
        filetypes=[
            ("Excel 파일", "*.xlsx"),
            ("모든 파일", "*.*")
        ],
        initialfile=default_filename,
        initialdir=initial_dir
    )
    
    root.destroy()  # tkinter 윈도우 제거
    
    # 파일이 저장되었으면 해당 디렉토리를 저장
    if save_path:
        directory = os.path.dirname(save_path)
        save_last_directory(directory)
    
    return save_path if save_path else None

def open_excel_file(file_path):
    """
    엑셀 파일을 운영체제 기본 프로그램으로 열기
    
    Args:
        file_path (str): 열려는 엑셀 파일 경로
    """
    try:
        if platform.system() == "Windows":
            os.startfile(file_path)
        elif platform.system() == "Darwin":  # macOS
            subprocess.run(["open", file_path])
        else:  # Linux
            subprocess.run(["xdg-open", file_path])
        
        print(f"엑셀 파일이 열렸습니다: {file_path}")
    except Exception as e:
        print(f"엑셀 파일을 여는 중 오류가 발생했습니다: {str(e)}")
        print(f"수동으로 파일을 열어주세요: {file_path}")

def process_pdf_to_excel(pdf_path, progress_window=None):
    """
    PDF 파일을 읽어서 엑셀로 변환하는 메인 처리 함수
    
    Args:
        pdf_path (str): PDF 파일 경로
        progress_window (ProgressWindow): 프로그래스바 객체
    """
    
    # 터미널 로그 수집용 리스트
    terminal_logs = []
    
    def log_and_print(message):
        """터미널에 출력하면서 동시에 로그에 저장하는 함수"""
        print(message)
        terminal_logs.append(message)
    
    if not os.path.exists(pdf_path):
        log_and_print(f"오류: '{pdf_path}' 파일을 찾을 수 없습니다.")
        if progress_window:
            progress_window.close()
        return
    
    try:
        if progress_window:
            progress_window.update_progress(5, "Opening PDF file...")
        
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            if total_pages == 0:
                print("PDF에 페이지가 없습니다.")
                if progress_window:
                    progress_window.close()
                return
            
            if progress_window:
                progress_window.update_progress(10, f"Analyzing PDF pages... (Total {total_pages} pages)")
            
            log_and_print(f"PDF 총 페이지 수: {total_pages}")
            all_extracted_data = []
            seq_no = None
            date = None
            global_test_counter = 0  # 전역 테스트 카운터
            
            # 첫 번째 페이지 처리
            first_page = pdf.pages[0]
            text = first_page.extract_text()
            
            if not text:
                log_and_print("첫 번째 페이지에서 텍스트를 추출할 수 없습니다.")
                if progress_window:
                    progress_window.close()
                return
            
            lines = text.split('\n')
            
            if progress_window:
                progress_window.update_progress(20, "Extracting data from first page...")
            
            # 디버깅용: 줄 번호와 내용 출력
            log_and_print("=" * 50)
            log_and_print("첫 번째 페이지 내용:")
            log_and_print("=" * 50)
            for i, line in enumerate(lines, 1):
                if line.strip():
                    log_and_print(f"줄 {i:3d}: {line}")
            log_and_print("=" * 50)
            
            # 첫 번째 페이지 데이터 추출
            base_seq_no, date, first_page_data, global_test_counter = extract_data_from_first_page(lines)
            all_extracted_data.extend(first_page_data)
            
            if progress_window:
                progress_window.update_progress(30, f"First page completed ({len(first_page_data)} data items)")
            
            log_and_print(f"\n첫 번째 페이지에서 추출된 데이터: {len(first_page_data)}개")
            
            # 두 번째 페이지부터 처리
            for page_num in range(1, total_pages):
                # 진행률 계산 (30%부터 60%까지)
                if progress_window and total_pages > 1:
                    progress = 30 + int((page_num / (total_pages - 1)) * 30)
                    progress_window.update_progress(progress, f"Processing page {page_num + 1}/{total_pages}...")
                
                page = pdf.pages[page_num]
                text = page.extract_text()
                
                if not text:
                    log_and_print(f"페이지 {page_num + 1}에서 텍스트를 추출할 수 없습니다.")
                    continue
                
                lines = text.split('\n')
                
                # 디버깅용: 줄 번호와 내용 출력
                log_and_print(f"\n[ 페이지 {page_num + 1} ]")
                log_and_print("-" * 30)
                for i, line in enumerate(lines, 1):
                    if line.strip():
                        log_and_print(f"줄 {i:3d}: {line}")
                
                # 두 번째 페이지부터의 데이터 추출
                page_seq_no, page_date, page_data, test_counter = extract_data_from_other_pages(lines, global_test_counter)
                global_test_counter = test_counter  # 전역 카운터 업데이트
                all_extracted_data.extend(page_data)
                
                log_and_print(f"페이지 {page_num + 1}에서 추출된 데이터: {len(page_data)}개")
                log_and_print(f"  - Seq No: {page_seq_no}, Date: {page_date}")
            
            if not all_extracted_data:
                log_and_print("추출할 데이터가 없습니다.")
                if progress_window:
                    progress_window.close()
                return
            
            if progress_window:
                progress_window.update_progress(60, "Organizing data...")
            
            log_and_print(f"\n전체 추출된 데이터:")
            log_and_print(f"Seq No: {seq_no}")
            log_and_print(f"Date: {date}")
            log_and_print(f"총 데이터 개수: {len(all_extracted_data)}")
            
            # 변수명 변경
            extracted_data = all_extracted_data
            
            if progress_window:
                progress_window.update_progress(70, "Selecting save location...")
            
            # 엑셀 파일 저장 위치 선택
            pdf_filename = os.path.basename(pdf_path)
            log_and_print("\n엑셀 파일 저장 위치를 선택해주세요...")
            output_path = select_save_location(pdf_filename)
            
            if not output_path:
                log_and_print("저장이 취소되었습니다.")
                if progress_window:
                    progress_window.close()
                return
            
            if progress_window:
                progress_window.update_progress(80, "Creating Excel file...")
            
            log_and_print(f"저장 위치: {output_path}")
            
            # PDF 줄별 데이터 수집
            pdf_lines = []
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                lines = text.split('\n')
                pdf_lines.append({
                    'page': page_num,
                    'lines': lines
                })
            
            # 엑셀 파일 생성 (터미널 로그 포함)
            create_excel_file(pdf_filename, extracted_data, output_path, terminal_logs, pdf_lines)
            
            if progress_window:
                progress_window.update_progress(95, "Opening Excel file...")
            
            # 엑셀 파일 자동 실행
            log_and_print("\n엑셀 파일을 열고 있습니다...")
            open_excel_file(output_path)
            
            if progress_window:
                progress_window.update_progress(100, "Completed!")
                time.sleep(1)  # 1초 대기 후 창 닫기
                progress_window.close()
            
    except Exception as e:
        log_and_print(f"PDF 처리 중 오류가 발생했습니다: {str(e)}")
        if progress_window:
            progress_window.close()

def select_pdf_file():
    """
    GUI로 PDF 파일을 선택하는 함수
    
    Returns:
        str: 선택된 PDF 파일의 경로, 취소시 None
    """
    if not TKINTER_AVAILABLE:
        return None
    
    # tkinter 윈도우 생성 (숨김)
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    
    # 마지막으로 사용한 디렉토리 불러오기
    initial_dir = load_last_directory()
    
    # 파일 선택 대화상자 열기
    pdf_path = filedialog.askopenfilename(
        title="PDF 파일을 선택하세요",
        filetypes=[
            ("PDF 파일", "*.pdf"),
            ("모든 파일", "*.*")
        ],
        initialdir=initial_dir  # 마지막 사용 디렉토리에서 시작
    )
    
    root.destroy()  # tkinter 윈도우 제거
    
    # 파일이 선택되었으면 해당 디렉토리를 저장
    if pdf_path:
        directory = os.path.dirname(pdf_path)
        save_last_directory(directory)
    
    return pdf_path if pdf_path else None

def run(pdf_path:str) -> str:
    """
    Entrypoint: converts PDF to Excel and returns output path
    Streamlit 환경에서 호출될 때는 파일 저장 대화상자를 표시하지 않고 임시 파일에 저장합니다.
    
    Args:
        pdf_path (str): PDF 파일 경로
        
    Returns:
        str: 생성된 Excel 파일 경로
    """
    # Streamlit 환경에서 실행 중인지 확인
    is_streamlit = 'streamlit' in sys.modules
    
    # 터미널 로그를 저장할 리스트
    terminal_logs = []
    
    def log_and_print(msg):
        terminal_logs.append(msg)
        print(msg)
    
    # 입력 파일 체크
    if not os.path.exists(pdf_path):
        log_and_print(f"오류: 파일을 찾을 수 없습니다: {pdf_path}")
        return None

    try:
        # PDF 열기
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            if total_pages == 0:
                log_and_print("PDF에 페이지가 없습니다.")
                return None

            # 첫 페이지 추출
            lines = pdf.pages[0].extract_text().split('\n')
            base_seq_no, date, first_page_data, global_test_counter = extract_data_from_first_page(lines)

            # 이후 페이지 추출
            for i, page in enumerate(pdf.pages[1:], start=1):
                lines = page.extract_text().split('\n')
                _, _, data, global_test_counter = extract_data_from_other_pages(lines, global_test_counter)
                first_page_data.extend(data)

        if not first_page_data:
            log_and_print("추출된 데이터가 없습니다.")
            return None

        # Streamlit 환경에서는 임시 파일에 저장
        if is_streamlit:
            import tempfile
            pdf_filename = os.path.basename(pdf_path)
            base_name = os.path.splitext(pdf_filename)[0]
            # 임시 파일 생성
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                output_path = tmp.name
            
            # 파일 이름을 PDF 파일명과 동일하게 설정 (확장자만 .xlsx로 변경)
            new_output_path = os.path.join(os.path.dirname(output_path), f"{base_name}.xlsx")
            if os.path.exists(new_output_path):
                try:
                    os.remove(new_output_path)  # 기존 파일이 있으면 삭제
                except:
                    pass
            try:
                os.rename(output_path, new_output_path)
                output_path = new_output_path
            except:
                # 이름 변경 실패 시 원래 임시 파일 경로 사용
                pass
        else:
            # 일반 환경에서는 사용자에게 저장 위치 선택 요청
            pdf_filename = os.path.basename(pdf_path)
            output_path = select_save_location(pdf_filename)
            if not output_path:
                return None

        # PDF 줄별 데이터 수집
        pdf_lines = []
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                lines = text.split('\n')
                pdf_lines.append({
                    'page': page_num,
                    'lines': lines
                })
        
        # 엑셀 생성
        create_excel_file(os.path.basename(pdf_path), first_page_data, output_path, terminal_logs, pdf_lines)
        return output_path

    except Exception as e:
        log_and_print(f"PDF 처리 중 오류 발생: {e}")
        return None

def main():
    """
    메인 함수: GUI로 PDF 파일을 선택받아 엑셀로 변환합니다.
    """
    
    if len(sys.argv) > 1:
        # 명령행 인수로 파일 경로가 제공된 경우
        pdf_path = sys.argv[1]
        print(f"명령행에서 제공된 파일: {pdf_path}")
    else:
        # GUI로 파일 선택
        print("PDF 파일 선택 창을 열고 있습니다...")
        pdf_path = select_pdf_file()
        
        if not pdf_path:
            print("파일 선택이 취소되었습니다.")
            return
            
        print(f"선택된 파일: {pdf_path}")
    
    # 프로그래스바 생성 및 표시
    progress_window = ProgressWindow()
    progress_window.show()
    
    # PDF를 엑셀로 변환
    process_pdf_to_excel(pdf_path, progress_window)

if __name__ == "__main__":
    main()
