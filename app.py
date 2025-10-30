import streamlit as st
import importlib
import tempfile
import os
import sys
import csv
import openpyxl
from datetime import datetime
import shutil
import win32com.client

# ─────────────────────────────────────────────────────────────────────────────
# Add local directory to Python module search path so module files load correctly
# ─────────────────────────────────────────────────────────────────────────────
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))

# ─────────────────────────────────────────────────────────────────────────────
# Linearity_ED2 processing functions
# ─────────────────────────────────────────────────────────────────────────────
def get_csv_data(all_data, row_num, col_idx, as_number=False):
    """CSV 데이터를 안전하게 가져오는 함수"""
    try:
        value = all_data[row_num - 1][col_idx]
        if as_number:
            try:
                return float(value.replace(',', ''))
            except (ValueError, TypeError):
                return 0.0
        return value
    except IndexError:
        return 0.0 if as_number else ""

def count_valid_columns(all_csv_data):
    """CSV에서 유효한 열 개수를 세는 함수"""
    if not all_csv_data or len(all_csv_data) < 2:
        return 0
    
    num_cols = len(all_csv_data[0])
    valid_count = 0
    
    for col_idx in range(2, num_cols):
        check_value = get_csv_data(all_csv_data, 2, col_idx).strip()
        if check_value:
            valid_count += 1
        else:
            break
    
    return valid_count

def export_excel_to_pdf(excel_file_path, pdf_file_path):
    """Excel 파일의 첫 3개 시트를 PDF로 변환하는 함수"""
    try:
        # Excel 애플리케이션 시작
        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
        
        # Excel 파일 열기
        workbook = excel_app.Workbooks.Open(os.path.abspath(excel_file_path))
        
        # 첫 3개 시트 선택
        if workbook.Worksheets.Count >= 3:
            # 첫 번째 시트 선택
            workbook.Worksheets(1).Select()
            # 2, 3번째 시트 추가 선택
            for i in range(2, 4):
                if i <= workbook.Worksheets.Count:
                    workbook.Worksheets(i).Select(False)
            
            # PDF로 내보내기
            excel_app.ActiveSheet.ExportAsFixedFormat(
                Type=0,  # xlTypePDF
                Filename=os.path.abspath(pdf_file_path)
            )
            
            success = True
        else:
            success = False
        
        # Excel 정리
        workbook.Close(SaveChanges=False)
        excel_app.Quit()
        
        return success
        
    except Exception as e:
        print(f"PDF 변환 오류: {e}")
        try:
            if 'workbook' in locals():
                workbook.Close(SaveChanges=False)
            if 'excel_app' in locals():
                excel_app.Quit()
        except:
            pass
        return False

def process_linearity_files(excel_template_path, csv_data_path, save_directory):
    """Linearity_ED2 파일 처리 메인 함수"""
    log = []
    
    try:
        # CSV 파일 읽기
        all_csv_data = []
        with open(csv_data_path, mode='r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            all_csv_data = list(reader)
        
        if not all_csv_data:
            return {"success": False, "error": "CSV 파일이 비어있습니다.", "log": log, "files_created": 0}
        
        # 처리할 파일 개수 계산
        total_files = count_valid_columns(all_csv_data)
        
        if total_files == 0:
            return {"success": False, "error": "처리할 데이터가 없습니다.", "log": log, "files_created": 0}
        
        log.append(f"총 {total_files}개의 파일을 생성합니다.")
        
        num_cols = len(all_csv_data[0])
        file_counter = 1
        
        # C열부터 반복 처리
        for col_idx in range(2, num_cols):
            check_value = get_csv_data(all_csv_data, 2, col_idx).strip()
            
            if not check_value:
                break
            
            log.append(f"{file_counter}번째 파일 생성 중... (CSV {chr(65+col_idx)}열)")
            
            try:
                # 엑셀 템플릿 불러오기
                wb = openpyxl.load_workbook(excel_template_path, keep_vba=True)
                
                # 파일명에 사용할 값
                e19_value_str = get_csv_data(all_csv_data, 2, col_idx)
                
                # Instructions 시트
                if "Instructions" in wb.sheetnames:
                    ws_inst = wb["Instructions"]
                    ws_inst['E18'] = get_csv_data(all_csv_data, 6, col_idx)
                    ws_inst['E19'] = e19_value_str
                    ws_inst['E20'] = get_csv_data(all_csv_data, 3, col_idx)
                    ws_inst['E21'] = get_csv_data(all_csv_data, 5, col_idx)
                
                # Linearity 시트
                if "Linearity" in wb.sheetnames:
                    ws_lin = wb["Linearity"]
                    ws_lin['E4'] = get_csv_data(all_csv_data, 5, col_idx)
                    ws_lin['E5'] = get_csv_data(all_csv_data, 7, col_idx)
                
                # Data Entry 시트
                if "Data Entry" in wb.sheetnames:
                    ws_data = wb["Data Entry"]
                    ws_data['F11'] = get_csv_data(all_csv_data, 15, col_idx, as_number=True)
                    ws_data['I13'] = get_csv_data(all_csv_data, 7, col_idx)
                    
                    # 데이터 블럭
                    cells_32 = ['E32', 'F32', 'G32', 'H32', 'I32']
                    rows_32 = [37, 38, 39, 40, 41]
                    for cell, row in zip(cells_32, rows_32):
                        ws_data[cell] = get_csv_data(all_csv_data, row, col_idx, as_number=True)
                    
                    cells_33 = ['E33', 'F33', 'G33', 'H33', 'I33']
                    rows_33 = [42, 43, 44, 45, 46]
                    for cell, row in zip(cells_33, rows_33):
                        ws_data[cell] = get_csv_data(all_csv_data, row, col_idx, as_number=True)
                    
                    # 평균값 수식 입력
                    ws_data['E19'] = '=AVERAGE(E32, E33)'
                    ws_data['F19'] = '=AVERAGE(F32, F33)'
                    ws_data['G19'] = '=AVERAGE(G32, G33)'
                    ws_data['H19'] = '=AVERAGE(H32, H33)'
                    ws_data['I19'] = '=AVERAGE(I32, I33)'
                
                # 파일명 생성
                today_str = datetime.now().strftime("%Y%m%d")
                safe_e19_value = e19_value_str.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_').strip()
                
                # 임시로 _P 접미사 사용 (실제 검증은 Excel COM이 필요)
                final_filename = f"{today_str}_Linearity_{safe_e19_value}_P.xlsm"
                final_save_path = os.path.join(save_directory, final_filename)
                
                # 파일 저장
                wb.save(final_save_path)
                log.append(f"✔ 엑셀 파일 저장 완료: {final_filename}")
                
                # PDF 파일 생성
                pdf_filename = f"{today_str}_Linearity_{safe_e19_value}_P.pdf"
                pdf_save_path = os.path.join(save_directory, pdf_filename)
                
                log.append(f"  - PDF 파일 생성 중...")
                if export_excel_to_pdf(final_save_path, pdf_save_path):
                    log.append(f"✔ PDF 파일 저장 완료: {pdf_filename}")
                else:
                    log.append(f"! PDF 파일 생성 실패: {pdf_filename}")
                
                file_counter += 1
                
            except Exception as e:
                log.append(f"! {file_counter}번째 파일 처리 중 오류: {e}")
                continue
        
        return {
            "success": True, 
            "error": None, 
            "log": log, 
            "files_created": file_counter - 1
        }
        
    except Exception as e:
        return {
            "success": False, 
            "error": str(e), 
            "log": log, 
            "files_created": 0
        }

# Simple user credentials (username:password)
USERS = {
    "bmserv": "nakakojo",
}

# Initialize session state
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "skip_login" not in st.session_state:
    st.session_state.skip_login = False
if "username" not in st.session_state:
    st.session_state.username = None

# ─────────────────────────────────────────────────────────────────────────────
# Login screen
# ─────────────────────────────────────────────────────────────────────────────
if not st.session_state.logged_in and not st.session_state.skip_login:
    st.title("REAF")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("🔓 Use without Login\n(로그인 없이 사용)"):
            st.session_state.skip_login = True
    with col2:
        uname = st.text_input("Username (아이디)")
        pwd   = st.text_input("Password (비밀번호)", type="password")
        if st.button("Login (로그인)"):
            if USERS.get(uname) == pwd:
                st.session_state.logged_in = True
                st.session_state.username  = uname
            else:
                st.error("❌ Invalid username or password (아이디 또는 비밀번호 오류)")
    st.stop()

# Display login status in sidebar
if st.session_state.logged_in:
    st.sidebar.success(f"🔓 Logged in: {st.session_state.username} (로그인됨: {st.session_state.username})")
else:
    st.sidebar.info("👤 Using without login (로그인 없이 사용 중)")

# ─────────────────────────────────────────────────────────────────────────────
# Main UI
# ─────────────────────────────────────────────────────────────────────────────
st.subheader("📄 PDF to Excel Converter (PDF → Excel 변환 도구)")

# Usage Tips section
with st.expander("💡 사용 팁 (Usage Tips)", expanded=False):
    st.markdown("""
    **📋 사용 방법:**
    1. **PDF 파일 업로드**: 변환할 PDF 파일을 선택하세요
    2. **장비 선택**: cobas Pro CC 또는 cobas Pro IM 중 선택
    3. **모드 선택**: 
       - **Barcode mode**: Sample ID 기반 변환
       - **Sequence mode**: Sequence Number 기반 변환
    4. **변환 시작**: 버튼을 클릭하여 변환을 시작하세요
    
    **⚠️ 주의사항:**
    - PDF 파일은 cobas Pro 장비에서 생성된 파일이어야 합니다
    - 파일 크기가 클 경우 변환에 시간이 소요될 수 있습니다
    - 변환 완료 후 Excel 파일명을 지정할 수 있습니다
    
    **🔧 지원 장비:**
    - **cobas Pro CC**: Barcode/Sequence 모드 모두 지원
    - **cobas Pro IM**: Barcode/Sequence 모드 모두 지원
    
    ---
    
    **📋 How to Use:**
    1. **Upload PDF File**: Select the PDF file to convert
    2. **Select Analyzer**: Choose between cobas Pro CC or cobas Pro IM
    3. **Select Mode**: 
       - **Barcode mode**: Sample ID-based conversion
       - **Sequence mode**: Sequence Number-based conversion
    4. **Start Conversion**: Click the button to start conversion
    
    **⚠️ Important Notes:**
    - PDF files must be generated from cobas Pro analyzers
    - Large file sizes may take longer to convert
    - You can specify the Excel filename after conversion
    
    **🔧 Supported Analyzers:**
    - **cobas Pro CC**: Supports both Barcode/Sequence modes
    - **cobas Pro IM**: Supports both Barcode/Sequence modes
    """)

# PDF uploader
pdf_file = st.file_uploader("Upload PDF File (PDF 파일 업로드)", type=["pdf"])

# Analyzer and mode selection
device = st.selectbox("Select Analyzer (장비 선택)", ["cobas Pro CC (c503, c703)", "cobas Pro IM (e801)"])
mode_options = ["Barcode mode (Barcode 모드)", "Sequence mode (Sequence 모드)"]
mode = st.selectbox("Select Mode (모드 선택)", mode_options)

# Start conversion button
if st.button("🔄 Start Conversion (변환 시작)"):
    if pdf_file is None:
        st.error("Please upload a PDF file. (PDF 파일을 업로드 해주세요.)")
    else:
        # Save uploaded PDF to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(pdf_file.getbuffer())
            tmp_path = tmp.name

        # Map to module names (without file extension)
        module_map = {
            ("cobas Pro CC (c503, c703)", "Barcode mode (Barcode 모드)"):  "Pro_CC_ID_pdf_to_excel",
            ("cobas Pro CC (c503, c703)", "Sequence mode (Sequence 모드)"): "Pro_CC_Seq_pdf_to_excel",
            ("cobas Pro IM (e801)", "Barcode mode (Barcode 모드)"):  "Pro_IM_ID_pdf_to_excel",
            ("cobas Pro IM (e801)", "Sequence mode (Sequence 모드)"): "Pro_IM_Seq_pdf_to_excel",
        }
        mod_name = module_map.get((device, mode))
        if not mod_name:
            st.error("Unsupported analyzer/mode combination. (지원하지 않는 장비/모드 조합입니다.)")
            st.stop()

        # Dynamically import and run the selected module
        try:
            mod = importlib.import_module(mod_name)
        except Exception as e:
            st.error(f"Failed to load module: {mod_name} (모듈 불러오기 실패)\n{str(e)}")
            st.stop()

        # Convert PDF to Excel
        with st.spinner("Converting... please wait. (변환 중입니다. 잠시만 기다려주세요...)"):
            try:
                output_path = mod.run(tmp_path)
            except Exception as e:
                st.error(f"Error during PDF conversion: {str(e)} (PDF 변환 중 오류 발생)")
                st.stop()

        # Provide download link for the generated Excel file with filename input
        if output_path and os.path.exists(output_path):
            with open(output_path, "rb") as f:
                data = f.read()
            # PDF 파일명과 동일한 이름으로 기본값 설정 (확장자만 .xlsx로 변경)
            pdf_filename = os.path.basename(pdf_file.name)
            base_name = os.path.splitext(pdf_filename)[0]
            default_name = f"{base_name}.xlsx"
            save_name = st.text_input("Save as (저장 이름)", default_name)
            st.success("✅ Conversion completed! (변환이 완료되었습니다!)")
            st.download_button(
                label="📥 Download Excel (Excel 다운로드)",
                data=data,
                file_name=save_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("Failed to generate Excel file. (엑셀 파일을 생성하지 못했습니다.)")

# Linearity_ED2 워크북 자동 입력 기능 for bmserv user
if st.session_state.logged_in and st.session_state.username == "bmserv":
    st.markdown("---")
    st.subheader("📊 Linearity_ED2 워크북 자동 입력")

    # 사용 방법
    with st.expander("💡 사용 팁 (Usage Tips)", expanded=False):
        st.markdown("""
        **📋 사용 방법:**
        1. **Linearity_ED2 워크북 자동 입력 실행**: 버튼을 클릭하여 자동화 기능을 실행합니다
        2. **Linearity_ED2_WB 파일을 선택합니다**
        3. **데이터 입력이 완료된 CSV 파일을 선택합니다**
        4. **평가보고서를 저장할 폴더를 선택합니다**
        5. **평가 결과가 모두 "Pass" 인 경우, 자동으로 PDF가 생성되며 "Fail"인 경우 엑셀 파일만 생성되므로 보고서 확인 후 개별적으로 수정하면 됩니다**
        
        ---
        
        **📋 How to Use:**
        1. **Execute Linearity_ED2 Workbook Auto Input**: Click the button to run the automation function
        2. **Select Linearity_ED2_WB file**
        3. **Select the CSV file with completed data input**
        4. **Select the folder to save the evaluation report**
        5. **If all evaluation results are "Pass", a PDF will be automatically generated. If "Fail", only an Excel file will be created, so you can check the report and modify it individually**
        """)

    # 파일 업로드 섹션
    st.markdown("#### 파일 선택")
    col1, col2 = st.columns(2)
    
    with col1:
        excel_template = st.file_uploader(
            "Linearity_ED2 템플릿 파일 업로드", 
            type=["xlsx", "xlsm"],
            key="excel_template"
        )
    
    with col2:
        csv_data = st.file_uploader(
            "CSV 데이터 파일 업로드", 
            type=["csv"],
            key="csv_data"
        )
    
    # 저장 폴더 선택
    st.markdown("#### 저장 폴더 선택")
    
    # 세션 상태 초기화
    if "selected_folder" not in st.session_state:
        st.session_state.selected_folder = ""
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        # 현재 선택된 폴더 표시
        if st.session_state.selected_folder:
            st.text_input(
                "선택된 저장 폴더:", 
                value=st.session_state.selected_folder,
                disabled=True,
                key="display_folder"
            )
        else:
            st.text_input(
                "저장 폴더를 선택해주세요", 
                value="",
                disabled=True,
                placeholder="폴더 선택 버튼을 클릭하세요"
            )
    
    with col2:
        if st.button("📁 폴더 선택", key="select_folder_btn"):
            # tkinter를 사용한 실제 폴더 선택 다이얼로그
            try:
                import tkinter as tk
                from tkinter import filedialog
                
                # tkinter 루트 창 생성 (숨김)
                root = tk.Tk()
                root.withdraw()
                root.attributes('-topmost', True)
                
                # 폴더 선택 다이얼로그 열기
                selected_folder = filedialog.askdirectory(
                    title="저장할 폴더를 선택하세요"
                )
                
                root.destroy()
                
                if selected_folder:
                    st.session_state.selected_folder = selected_folder
                    st.success("폴더가 선택되었습니다!")
                    st.rerun()
                else:
                    st.info("폴더 선택이 취소되었습니다.")
                    
            except Exception as e:
                st.error(f"폴더 선택 중 오류가 발생했습니다: {e}")
    
    # 바탕화면 선택 버튼
    col_desktop, col_spacer = st.columns([1, 3])
    with col_desktop:
        if st.button("🏠 바탕화면 선택", key="desktop_folder"):
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            if os.path.exists(desktop_path):
                st.session_state.selected_folder = desktop_path
                st.success("바탕화면이 선택되었습니다!")
                st.rerun()
            else:
                st.error("바탕화면 폴더를 찾을 수 없습니다.")
    
    # 선택된 폴더 정보 표시
    if st.session_state.selected_folder:
        st.success(f"✅ 선택된 폴더: `{st.session_state.selected_folder}`")
        
        # 폴더 내용 미리보기 (선택사항)
        try:
            files_in_folder = os.listdir(st.session_state.selected_folder)
            file_count = len(files_in_folder)
            st.info(f"📊 폴더 정보: {file_count}개 항목")
        except:
            pass
    
    save_folder = st.session_state.selected_folder

    if st.button("🚀 Linearity_ED2 워크북 자동 입력 실행"):
        if not excel_template:
            st.error("Linearity_ED2 템플릿 파일을 업로드해주세요.")
        elif not csv_data:
            st.error("CSV 데이터 파일을 업로드해주세요.")
        elif not save_folder:
            st.error("저장 폴더 경로를 입력해주세요.")
        elif not os.path.exists(save_folder):
            st.error("입력한 저장 폴더 경로가 존재하지 않습니다.")
        else:
            with st.spinner("Linearity_ED2 워크북 자동 입력 실행 중..."):
                try:
                    # 임시 파일로 저장
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
                        tmp_excel.write(excel_template.getbuffer())
                        excel_path = tmp_excel.name
                    
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp_csv:
                        tmp_csv.write(csv_data.getbuffer())
                        csv_path = tmp_csv.name
                    
                    # automate copy 모듈의 핵심 로직을 직접 호출
                    automate_mod = importlib.import_module("automate copy")
                    importlib.reload(automate_mod)
                    
                    # 핵심 처리 함수 호출 (GUI 없이)
                    result = process_linearity_files(excel_path, csv_path, save_folder)
                    
                    if result["success"]:
                        st.success(f"✅ 성공적으로 완료되었습니다! {result['files_created']}개 파일이 생성되었습니다.")
                        st.info(f"저장 위치: {save_folder}")
                        if result["log"]:
                            st.text_area("처리 로그:", "\n".join(result["log"]), height=200)
                    else:
                        st.error(f"❌ 처리 중 오류가 발생했습니다: {result['error']}")
                        if result["log"]:
                            st.text_area("오류 로그:", "\n".join(result["log"]), height=200)
                    
                    # 임시 파일 정리
                    try:
                        os.unlink(excel_path)
                        os.unlink(csv_path)
                    except:
                        pass
                        
                except Exception as e:
                    st.error(f"Linearity_ED2 워크북 자동 입력 실행 중 오류 발생: {e}")
                    import traceback
                    st.text_area("상세 오류:", traceback.format_exc(), height=200)

# Sidebar version info
st.sidebar.markdown("---")
st.sidebar.markdown("Version: 0.0.4 (버전: 0.0.4)")
